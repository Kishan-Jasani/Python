#Write Only (‘w’): 
# --> Open the file for writing. For an existing file, the data is truncated and over-written.

#Read Only (‘r’): 
# --> Open the file for reading.

#Write and Read (‘w+’): 
# --> Open the file for reading and writing. For an existing file, data is truncated and over-written.

#Append Only (‘a’): 
# --> Open the file for writing. The data being written will be inserted at the end, after the existing data.

#Append and Read (‘a+’): 
# --> Open the file for reading and writing. The data being written will be inserted at the end, after the existing data.


#open(file_path,operation)
# --> This method is used to open the file for the specified operation. The operation can either be r,w,a for read, write and append.
open_file = open("D:\File_Handling.txt","w")


#write()
# --> This method is used to write a string to a file, if file is present. If not, it creates the file and writes the string into it.
open_file.write("Hello")



#close()
# --> This method is used to close a file which is already open.
open_file.close()



#read()
# --> This method is used to read all the contents of a file into a string.
try:
    read_file=open("D:\File_Handling.txt","r")
    text=read_file.read()
    print(text)
    read_file.write(",Good Morning")    #Throw exception here because file has only read access.
    read_file.close()
except:
    print("Error occurred")
    if read_file.closed:
        print("File is closed")
    else:
        print("File is open")
        
        


#Python programming using Excel Files
# --> https://lex.infosysapps.com/en/viewer/web-module/lex_auth_013003669947932672118?collectionId=lex_auth_01278069079858380860&collectionType=Course&viewMode=START

import openpyxl as xl
wb=xl.load_workbook(r'Employee.xlsx')
print('Sheet names:',wb.sheetnames)
ws=wb['Emp']
print('Type of ws:',type(ws))

cell1=ws.cell(1,1)


cell2=ws['A2']

cell1.value
cell2.value

ws.max_row
ws.max_column

cells=ws['A1':'C4']
for row in cells:
    for cell in row:
        print(cell.value,end=' ')
    print()



wb=xl.load_workbook(r'Employee.xlsx')
ws=wb['Emp']
cells=tuple()
for row_num in range(2,ws.max_row+1):
    if ws.cell(row_num,1).value==3:  
        cells=ws[row_num]
        break
if cells:
    for cell in cells:
        print(cell.value,end=' ')
else:
    print('Employee record not found')






#Python Programming using CSV Files

import csv
with open(r'C:\Users\brahmeswara.d\Desktop\Customer.csv','r') as csvfile:
    reader=csv.reader(csvfile)
    reader=csv.DictReader(csvfile)
    for record in reader:
        print(record)


with open(r'C:\Users\brahmeswara.d\Desktop\Employees.csv','w') as csvfile:
    writer=csv.writer(csvfile)
    writer.writerow(['EmpId','Emp Name','Salary'])
    writer.writerow([1,'Ramesh',22001.00])
    writer.writerow([2,'Rakesh',26501.00])
   
   

with open(r'C:\Users\brahmeswara.d\Desktop\Employees.csv','w',newline='') as csvfile:
    fields=['EmpId','Emp Name','Salary']
    writer=csv.DictWriter(csvfile,fields)
    writer.writeheader()
    writer.writerow({'EmpId':1,'Emp Name':'Ramesh','Salary':22001.00})
    writer.writerow({'EmpId':2,'Emp Name':'Rakesh','Salary':26501.00})
   


